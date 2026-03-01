import os
import openpyxl
from django.core.management.base import BaseCommand
from django.core.files import File
from main.models import User, Category, Unit, Manufacturer, Supplier, Product, Order, OrderItem
from django.contrib.auth.hashers import make_password
from datetime import datetime

class Command(BaseCommand):
    help = 'Импорт данных из Excel файлов'

    def handle(self, *args, **options):
        import_path = 'import/'
        self.import_users(os.path.join(import_path, 'user_import.xlsx'))
        self.import_dictionaries_and_products(os.path.join(import_path, 'Tovar.xlsx'))
        self.import_orders(os.path.join(import_path, 'Заказ_import.xlsx'))
        self.import_pickup_points(os.path.join(import_path, 'Пункты выдачи_import.xlsx'))
        self.stdout.write(self.style.SUCCESS('Импорт завершён'))

    def import_users(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        role_map = {
            'Администратор': 'admin',
            'Менеджер': 'manager',
            'Клиент': 'client',
            'Авторизованный клиент': 'client'
        }
        for row in sheet.iter_rows(min_row=2, values_only=True):
            role_raw, fio, username, password = row[:4]
            if not username:
                continue
            fio_parts = fio.split() if fio else ['', '']
            last_name = fio_parts[0] if len(fio_parts) > 0 else ''
            first_name = fio_parts[1] if len(fio_parts) > 1 else ''
            role = role_map.get(role_raw, 'client')
            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    'password': make_password(password),
                    'last_name': last_name,
                    'first_name': first_name,
                    'role': role
                }
            )
            if created:
                self.stdout.write(f'Пользователь {username} создан')

    def import_dictionaries_and_products(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            article, name, unit_name, price, supplier_name, manufacturer_name, category_name, discount, quantity, description, photo = row
            if not article or not name:
                continue
            unit, _ = Unit.objects.get_or_create(name=unit_name)
            category, _ = Category.objects.get_or_create(name=category_name)
            manufacturer, _ = Manufacturer.objects.get_or_create(name=manufacturer_name)
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
            product, created = Product.objects.update_or_create(
                article=article,
                defaults={
                    'name': name,
                    'category': category,
                    'description': description or '',
                    'manufacturer': manufacturer,
                    'supplier': supplier,
                    'price': price,
                    'unit': unit,
                    'quantity_in_stock': quantity or 0,
                    'discount': discount or 0,
                }
            )
            if photo and isinstance(photo, str) and photo.lower().endswith(('.jpg', '.png')):
                photo_path = os.path.join('import', photo)
                if os.path.exists(photo_path):
                    with open(photo_path, 'rb') as f:
                        product.image.save(photo, File(f), save=True)
            if created:
                self.stdout.write(f'Товар {name} создан')

        def import_orders(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        status_map = {
            'Новый': 'new',
            'В обработке': 'processing',
            'Выполнен': 'completed',
            'Отменён': 'cancelled'
        }
        for row in sheet.iter_rows(min_row=2, values_only=True):
            order_number, article_order, delivery_date, address, fio_client, pickup_code, status_raw, *_ = row
            if not order_number:
                continue
            try:
                delivery_date_parsed = datetime.strptime(delivery_date, '%d.%m.%Y') if delivery_date else None
            except:
                delivery_date_parsed = None
            status = status_map.get(status_raw, 'new')
            user = None
            if fio_client:
                fio_client = str(fio_client)
                fio_parts = fio_client.split()
                last_name = fio_parts[0] if len(fio_parts) > 0 else ''
                first_name = fio_parts[1] if len(fio_parts) > 1 else ''
                user = User.objects.filter(last_name=last_name, first_name=first_name).first()
            order, created = Order.objects.update_or_create(
                order_number=order_number,
                defaults={
                    'status': status,
                    'pickup_point': address or '',
                    'issue_date': delivery_date_parsed,
                    'user': user
                }
            )
            if created:
                self.stdout.write(f'Заказ {order_number} создан')
            else:
                self.stdout.write(f'Заказ {order_number} обновлён')

    def import_pickup_points(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            address = row[0]
            if address and isinstance(address, str):
                pass
                